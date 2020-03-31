import pandas as pd
from flask import Flask, request, jsonify, json, render_template
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
import os
from flask_mail import Mail, Message
import config
import time
from datetime import timedelta, datetime
from flask_cors import CORS, cross_origin

def append_df_to_excel(filename, df):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    """
    from openpyxl import load_workbook

    import pandas as pd

    writer = pd.ExcelWriter(filename, engine='openpyxl')
    # try to open an existing workbook
    writer.book = load_workbook(filename)
    sheet_name = 'Sheet1'
    startrow = writer.book[sheet_name].max_row

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, index=False, header=None)

    # save the workbook
    writer.save()


def generate_order_number(filepath='data/ORDERS.xlsx'):
    df = pd.read_excel(filepath)
    alph = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
            'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    temp = str(100000 + len(df['ORDER ID']) + np.random.randint(1, 200000) + np.random.randint(1, 200000))
    return (np.random.choice(alph) + np.random.choice(alph) + temp)

application = app = Flask(__name__)

app.config['MAIL_SERVER'] = config.MAIL_SERVER
app.config['MAIL_PORT'] = config.MAIL_PORT
app.config['MAIL_USE_TLS'] = config.MAIL_USE_TLS
app.config['MAIL_USERNAME'] = config.MAIL_USERNAME
app.config['MAIL_PASSWORD'] = config.MAIL_PASSWORD




mail = Mail(app)

cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

#msg = Message('test subject', sender=config.MAIL_USERNAME, recipients = ['tanaya.invictus@gmail.com',
#                                                                         'randhir66@gmail.com'])
#msg.body = 'hi'

#food = pd.read_excel('data/inventory.xlsx', index_col='Item Code', sheet_name='Food')
#toiletry = pd.read_excel('data/inventory.xlsx', index_col='Item Code', sheet_name='Toiletry')
#cart = pd.DataFrame(columns=['Item', 'Quantity', 'Price'])
dropzone = pd.read_excel('data/dropzone.xlsx',usecols=['Drop Point', 'Drop Time'])
inventory = pd.read_excel('data/inventory.xlsx')



@app.route('/')
def home():
    return render_template('index.html')


#this sends inventory json data
@app.route('/test',methods=['POST','GET'])
@cross_origin()
def test():
    '''
    display = []
    for i in inventory['Item Name']:
        j = min(inventory[inventory['Item Name'] == i]['Item Qty'].values[0],
                inventory[inventory['Item Name'] == i]['Tab'].values[0])
        if j == 0:
            j = 'Item Out of Stock'
        k = inventory[inventory['Item Name'] == i]['Category'].values[0]
        l = inventory[inventory['Item Name'] == i]['Item Price'].values[0]
        display.append({'Item Name': i, 'Tab': str(j), 'Category':k, 'Price':str(l)})

    dic = {}
    for i in inventory['Category'].unique():
        dic[i] = []
    for key in dic.keys():
        temp = inventory[inventory['Category'] == key][['Item Name', 'Item Price', 'Tab']]
        for i in temp['Item Name']:
            #j = min(temp[temp['Item Name'] == i]['Item Qty'].values[0],
            #        temp[temp['Item Name'] == i]['Tab'].values[0])
            j = temp[temp['Item Name'] == i]['Tab'].values[0]
            if j == 0:
                j = 'Item Out of Stock'
            l = temp[temp['Item Name'] == i]['Item Price'].values[0]
            #dic[key].append({'Item Name': i, 'Tab': str(j), 'Price': str(l)})
            dic[key].append({i: [str(l), str(j)]})
    '''
    dic = {}
    for i in inventory['Category'].unique():
        display = []
        temp = inventory[inventory['Category'] == i]
        dic[i] = display
        for j in temp['Item Name']:
            k = list(np.arange(1,temp[temp['Item Name'] == j]['Tab'].values[0]+1))
            k = [str(i) for i in k]
            #if j == 0:
            #    j = 'Item Out of Stock'
            l = temp[temp['Item Name'] == j]['Item Price'].values[0]
            m = temp[temp['Item Name'] == j]['Item Code'].values[0]
            display.append({'Item Name': j, 'Tab': k, 'Price': str(l), 'Item Code': str(m)})
    return jsonify(dic)




#this sends dropzone and time json data
@app.route('/trial', methods=['POST','GET'])
@cross_origin()
def trial():
    dic = {}
    for i in dropzone['Drop Point'].unique():
        dic[i] = []
        for j in dropzone[dropzone['Drop Point'] == i]['Drop Time'].values:
            dic[i].append(j)
    return jsonify(dic)



#this receives the order json data
@app.route('/submit-order',methods=['GET','POST'])
@cross_origin()
def order_submit():
    #data = json.loads(request.data)
    #request.get(data)
    #print("//////////////////////////")
    #print(data)

    data = {"Name": "Karan", "Zone": "CANT A!11:30 AM",
            'Mobile': '6078820136',
            "Order": [{'item': 'Sugar', 'quantity': '3', 'category': 'food', 'price': '20', 'code': '1123'},
                      {'item': 'Toilet Paper', 'quantity': '1', 'category': 'toiletry', 'price': '30', 'code': '1123'},
                      {'item': 'Salt', 'quantity': '1', 'category': 'food', 'price': '10', 'code': '1123'},
                      {'item': 'Potatoes', 'quantity': '5', 'category': 'food', 'price': '25', 'code': '1123'}],
            "Feedback": "Hey"}

    order_number = generate_order_number()


    today = datetime.today().strftime("%d-%m-%Y")
    #tomorrow = datetime.today() + timedelta(1)
    #tomorrow = tomorrow.strftime("%d-%m-%Y")

    df2 = pd.DataFrame(data['Order'])
    df2['quantity'] = df2['quantity'].apply(len)
    #df2 = df2.drop(columns=['category'])

    tc = round(((df2['price'].apply(float)) * (df2['quantity'].apply(int))).sum(),2) + 3

    z_t = data['Zone'].split('!')
    df1 = pd.DataFrame({'Order Number': str(order_number),
                        'Name': data['Name'],
                        'Delivery Zone & Time': str(z_t[0]+' '+z_t[1]),
                        'Order Date' : today,
                        'Phone Number': data['Mobile'],
                        'Total Cost': str(tc)}, index=[1])

    #order = {'ORDER ID': str(order_number), 'ORDER': [data['Order']],
    #         'NAME': data['Name'], 'ZONE': data['Zone'], 'TIME': data['Time'], 'TOTAL COST': tc,
    #         'STATUS': 'Received'}

    #df3 = pd.DataFrame(order)

    #append_df_to_excel('data/ORDERS.xlsx',df3)

    #for i in df2[['item', 'quantity']].iterrows():
    #    temp = inventory['Item Name'] == i[1].values[0]
    #    qty = inventory[temp]['Item Qty']
    #    inventory.loc[(temp), 'Item Qty'] = int(qty) - int(i[1].values[1])
    #    inventory.to_excel('data/inventory_combined.xlsx')

    filename  = "orders/OrderNumber{}.pdf".format(order_number)
    fig, (ax1, ax2, ax3) = plt.subplots(nrows=3, ncols=1, figsize=(12, 8))

    ax1.axis('tight')
    ax1.axis('off')
    ax1.set_title('Bill (Final Amount has been rounded and includes a packing charge of 3₹)')
    the_table1 = ax1.table(cellText=df1.values, colLabels=df1.columns, loc='center')
    ax2.set_title("Breakdown")
    ax2.axis('tight')
    ax2.axis('off')
    the_table2 = ax2.table(cellText=df2.values, colLabels=df2.columns, loc='center')

    if data['Feedback']:
        df3 = pd.DataFrame({'Feedback or Request: ': data['Feedback']}, index=[1])

        ax3.set_title("Feedback or Request")
        ax3.axis('tight')
        ax3.axis('off')
        the_table3 = ax3.table(cellText= df3.values, colLabels=df3.columns, loc='center')

    pp = PdfPages(filename)
    pp.savefig(fig, bbox_inches='tight')
    pp.close()
    subject = 'New Order # {} is placed'.format(order_number)
    msg = Message(subject, sender='udhampurcanteen@gmail.com', recipients=['udhampurcanteen@gmail.com'])#,'karthik99th.tnk@gmail.com','karan.maheshwari14@gmail.com'])
    msg.body = "This is an automated email. Check the attachment for details of the new order"
    with app.open_resource(filename) as fp:
        msg.attach('Order Number {}'.format(order_number), 'application/pdf', fp.read())
    mail.send(msg)
    
    return "Success"



if __name__ == '__main__':
    app.run(debug=True)
